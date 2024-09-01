import os.path
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from base64 import urlsafe_b64decode
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime, timedelta
from time import sleep


# If modifying these scopes, delete the file token.json.
SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]

s1 = "<p class=\"paragraph\" data-qa=\"mail__text\"  style=\"margin-bottom: 0; margin-top: 0;\"><a style=\"color: #0096ff;text-decoration: none;\" target=\"_blank\" href=\""

def main():
  print("За сколько дней обработать почту?")
  print("0 - за сегодняшний день")
  print("1 - за вчерашний день")
  print("2 - за вчерашний и позавчерашний день")
  print("3 - за последние три дня (но не за сегодня)")
  print("Можно ввести больше 3, но программа может обработать только 500 писем.")
  mode = int(input("Введите число: "))



  creds = None
  # The file token.json stores the user's access and refresh tokens, and is
  # created automatically when the authorization flow completes for the first
  # time.
  if os.path.exists("token.json"):
    creds = Credentials.from_authorized_user_file("token.json", SCOPES)
  # If there are no (valid) credentials available, let the user log in.
  if not creds or not creds.valid:
    if creds and creds.expired and creds.refresh_token:
      creds.refresh(Request())
    else:
      flow = InstalledAppFlow.from_client_secrets_file(
          "credentials.json", SCOPES
      )
      creds = flow.run_local_server(port=0)
    # Save the credentials for the next run
    with open("token.json", "w") as token:
      token.write(creds.to_json())

  # Инициализируем пустой список для хранения информации о вакансиях
  vacancies = []

  try:
    # Call the Gmail API
    service = build("gmail", "v1", credentials=creds)
    if mode == 0:
      date1 = datetime.now()
      date1 = f"{date1.year}/{date1.month}/{date1.day}"
      date2 = datetime.now() + timedelta(days=1)
      date2 = f"{date2.year}/{date2.month}/{date2.day}"
    else:
      date1 = datetime.now() - timedelta(days=mode)
      date1 = f"{date1.year}/{date1.month}/{date1.day}"
      date2 = datetime.now()
      date2 = f"{date2.year}/{date2.month}/{date2.day}"
    results = service.users().messages().list(userId="me", maxResults=500, q = f"Вакансии по подписке after:{date1} before:{date2}").execute()
    messages = results.get("messages", [])

    if not messages:
      print("Писем не найдено.")
      return
    
    for message in messages:
      # Получение сообщения
      msg = service.users().messages().get(userId="me", id=message["id"]).execute() 

      # Дешифровка сообщения
      msg_str = urlsafe_b64decode(msg['payload']['body']['data'].encode('UTF8'))
      html_content = msg_str.decode('UTF8')
      
      # Ищем дату в тексте письма
      date_pattern = html_content.find("sent_date=")
      vacancy_date = html_content[date_pattern+18:date_pattern+20] + '.' + html_content[date_pattern+15:date_pattern+17] + '.' + html_content[date_pattern+12:date_pattern+14]

      # Начало парсинга
      soup = BeautifulSoup(html_content, 'html.parser')

      # Поиск всех тегов <a> с атрибутом href (ссылка) и парсинг их для поиска вакансий
      for link in soup.find_all('a', href=True):
          href = link['href']
          
          # Проверяем, что ссылка ведет на hh.ru/vacancy, что указывает на страницу вакансии
          if 'hh.ru/vacancy' in href:
              title = link.get_text(strip=True)  # Название вакансии, текст внутри <a>
              
              # Ищем следующий элемент после ссылки на вакансию
              next_element = link.find_parent('tr').find_next_sibling('tr')
              
              if next_element:
                # Проверяем следующий элемент и ищем информацию о компании и зарплате
                text = next_element.get_text(strip=True)
                    
                if "₽" in text:  # Если найден символ рубля, это зарплата
                    next_element = next_element.find_next_sibling('tr')
                    text = next_element.get_text(strip=True)
                    company = text
                else:
                    company = text  # Иначе это название компании
                
                # Добавляем найденную информацию в список вакансий
                vacancies.append({
                    'date': vacancy_date,
                    'url': href,
                    'title': title,
                    'company': company
                })

    vacancies.reverse()

    # Создание Excel файла
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vacancies"

    # Запись заголовков
    headers = ['Дата', 'Компания', 'Название вакансии']
    sheet.append(headers)

    # Запись данных в файл Excel
    for row_num, vacancy in enumerate(vacancies, start=2):  # начинаем с 2, так как 1 строка - заголовки
        # Запись даты вакансии
        sheet.cell(row=row_num, column=1, value=vacancy['date'])
        # Добавляем компанию
        sheet.cell(row=row_num, column=2, value=vacancy['company'])

        # Запись названия вакансии
        job_cell = sheet.cell(row=row_num, column=3)
        job_cell.value = vacancy['title']
        job_cell.hyperlink = vacancy['url']  # добавляем гиперссылку
        job_cell.style = "Hyperlink"  # применяем стиль гиперссылки

    # Сохранение Excel файла
    excel_file_path = 'вакансии.xlsx'
    workbook.save(excel_file_path)

    print(f"Данные успешно записаны в файл {excel_file_path}")

  except HttpError as error:
    # TODO(developer) - Handle errors from gmail API.
    print(f"An error occurred: {error}")


if __name__ == "__main__":
  main()
  sleep(1)